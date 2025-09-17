import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# 设置Excel文件的绝对路径，请根据实际情况修改
excel_path = r'D:\Big Files\学习\港城莞\Sem A Deadline Schedule.xlsx'


def update_cs5489_deadlines():
    # 加载工作簿
    wb = openpyxl.load_workbook(excel_path)

    # 获取第一个sheet "Deadline Schedule"
    sheet = wb["Deadline Schedule"]

    # 定义要添加的截止时间：周次和任务描述
    deadlines = [
        (4, "作业1截止"),
        (6, "作业2截止"),
        (8, "作业3截止"),
        (10, "作业4截止"),
        (13, "课程项目截止")
    ]

    # Friday列是N列
    col_letter = "N"

    # 遍历周次，找到对应行并添加任务
    for week, task in deadlines:
        for row in range(1, sheet.max_row + 1):
            if sheet[f"A{row}"].value == f"Week {week}":
                sheet[f"{col_letter}{row}"] = task
                break

    # 检查是否已存在"CS5489" sheet，存在则删除
    if "CS5489" in wb.sheetnames:
        del wb["CS5489"]

    # 创建新sheet "CS5489"
    new_sheet = wb.create_sheet("CS5489")

    # 设置表头
    headers = ["任务名称", "任务比重", "任务需求"]
    for col_num, header in enumerate(headers, 1):
        cell = new_sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 任务数据
    tasks = [
        ("教程练习", 0.1, "共8个教程练习，每个练习旨在加深对机器学习算法的理解，需在教程结束后一周内提交。"),
        ("作业1", 0.075, "第一次作业，涉及解决机器学习中的数学问题。提交截止时间为第4周。"),
        ("作业2", 0.075, "第二次作业，涉及解决机器学习中的数学问题。提交截止时间为第6周。"),
        ("作业3", 0.075, "第三次作业，涉及解决机器学习中的数学问题。提交截止时间为第8周。"),
        ("作业4", 0.075, "第四次作业，涉及解决机器学习中的数学问题。提交截止时间为第10周。"),
        ("课程项目", 0.3,
         "应用机器学习解决一个实际问题。项目需包括报告、代码实现和可选演示。最多3人一组。提交截止时间为第13周。"),
        ("期末考试", 0.3,
         "期末考试覆盖所有课程内容，包括选择题、计算题和理论题。考试时间为2小时，具体时间由大学安排。必须通过考试（至少达到最高考试分数的30%）才能通过课程。")
    ]

    # 写入任务数据
    for row_num, (name, weight, requirement) in enumerate(tasks, 2):
        new_sheet.cell(row=row_num, column=1, value=name)
        new_sheet.cell(row=row_num, column=2, value=weight)
        new_sheet.cell(row=row_num, column=3, value=requirement)

    # 调整列宽
    for col in new_sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        new_sheet.column_dimensions[col_letter].width = max_length + 2

    # 保存工作簿
    wb.save(excel_path)
    print("Excel文件更新完成！")


if __name__ == "__main__":
    update_cs5489_deadlines()